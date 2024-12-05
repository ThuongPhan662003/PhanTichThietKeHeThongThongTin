import os
from flask import (
    Blueprint,
    Response,
    render_template,
    request,
    flash,
    jsonify,
    redirect,
    url_for,
)
from flask_login import login_required, current_user
import openpyxl
import pandas as pd
from prompt_toolkit import HTML

from website import db
from sqlalchemy import func, cast, Date, text
from datetime import datetime
from website.role import role_required
from website.models import *
from website.models import NguoiDung, NhomNguoiDung
from flask import Response, render_template, request

import io
from flask import Response, request
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
import io
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors


# from website.webforms import
import json

report = Blueprint("report", __name__)


def convert_date_format(date_str):
    """Chuyển đổi ngày từ định dạng dd-mm-yyyy sang yyyy-mm-dd."""
    if date_str:  # Kiểm tra nếu date_str không phải None hoặc rỗng
        return datetime.strptime(date_str, "%d-%m-%Y").strftime("%Y-%m-%d")
    return None  # Trả về None nếu không có ngày


@report.route("/", methods=["GET"])
@role_required(["Quản lý"])
def report_list():
    """
    Hiển thị danh sách các báo cáo
    """
    # Danh sách các báo cáo với tiêu đề và route liên quan
    reports = [
        {
            "name": "Doanh thu theo tháng",
            "route": url_for("report.report_page_revenue"),
        },
        {
            "name": "Top món ăn bán chạy",
            "route": url_for("report.report_topdishes_page"),
        },
    ]

    return render_template("admin/report/report_list.html", reports=reports)


def get_dishes_data(start_date, end_date, limit_number=10):
    """
    Gọi stored procedure `GetTopSellingDishes` để lấy dữ liệu món ăn bán chạy.
    Nếu tham số ngày không được truyền vào, nó sẽ là None.
    """
    # Chuyển đổi ngày từ định dạng dd-mm-yyyy sang yyyy-mm-dd
    start_date = convert_date_format(start_date)
    end_date = convert_date_format(end_date)

    # Thực thi stored procedure với các tham số đã chuyển đổi
    result = db.session.execute(
        text(
            """
            CALL GetTopSellingDishes(:start_date, :end_date, :limit_number)
            """
        ),
        {
            "start_date": start_date,
            "end_date": end_date,
            "limit_number": limit_number,
        },
    )

    rows = result.fetchall()  # Đọc hết dữ liệu từ câu lệnh

    # Kiểm tra dữ liệu
    print(rows)  # In ra kết quả để kiểm tra cấu trúc dữ liệu (rows)

    # Chuyển đổi kết quả thành danh sách dict
    if rows:
        columns = result.keys()  # Lấy tên các cột từ kết quả SQL
        return [
            dict(zip(columns, row)) for row in rows
        ]  # Kết hợp cột và dữ liệu thành dict

    return []  # Trả về danh sách rỗng nếu không có dữ liệu

@report.route("/dishes", methods=["GET", "POST"])
@role_required(["Quản lý"])
def report_topdishes_page():
    if request.method == "POST":
        # Lấy thông tin từ form

        start_date = request.form.get("start_date")
        end_date = request.form.get("end_date")
        limit_number = int(request.form.get("limit_number", 10))
        if not start_date or not end_date or not limit_number:
            flash("Vui lòng nhập đầy đủ thông tin", "danger")
            return render_template("/admin/report/form_topdishes.html")
        try:
            # Chuyển đổi start_date và end_date sang định dạng datetime
            start_date_obj = datetime.strptime(start_date, "%d-%m-%Y")
            end_date_obj = datetime.strptime(end_date, "%d-%m-%Y")
        except ValueError:
            flash("Ngày tháng không hợp lệ. Vui lòng nhập theo định dạng dd-mm-yyyy.", "danger")
            return render_template("/admin/report/form_topdishes.html")
        if start_date_obj > end_date_obj:
            flash("Ngày bắt đầu không thể lớn hơn ngày kết thúc", "danger")
            return render_template("/admin/report/form_topdishes.html")
        try:
            limit_number = int(limit_number)
            if limit_number <= 0:
                flash("Số món ăn cần lấy phải lớn hơn 0.", "danger")
                return render_template("/admin/report/form_topdishes.html")
        except ValueError:
            flash("Giới hạn món ăn phải là một số nguyên hợp lệ.", "danger")
            return render_template("/admin/report/form_topdishes.html")
        # Gọi hàm lấy dữ liệu
        dishes = get_dishes_data(start_date, end_date, limit_number)
        print(dishes)
        return render_template(
            "/admin/report/report_topdishes.html",
            dishes=dishes,
            start_date=start_date,
            end_date=end_date,
        )

    return render_template("/admin/report/form_topdishes.html")
@report.route("/excel_topdishes")
@role_required(["Quản lý"])
def report_excel_dishes():
    # Lấy tham số từ URL (GET request)
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    limit_number = request.args.get("limit_number", default=10, type=int)
    employee = NhanVien.query.filter(NhanVien.idNguoiDung == current_user.get_id()).first().getHoTenNV()

    printed_by = employee  # Example printed_by, this should be dynamic based on the user
    printed_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Current date and time

    # Lấy dữ liệu món ăn
    dishes = get_dishes_data(start_date, end_date, limit_number)

    # Tạo DataFrame từ kết quả
    df = pd.DataFrame(dishes)

    # Xuất file Excel từ DataFrame
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Top Dishes Report")

        # Accessing the workbook and sheet
        workbook = writer.book
        worksheet = workbook["Top Dishes Report"]

        # Adding Printed by and Printed Date info
        worksheet["A1"] = "Printed by:"
        worksheet["B1"] = printed_by
        worksheet["A2"] = "Printed on:"
        worksheet["B2"] = printed_date

        # Format header
        header = worksheet[1]
        for cell in header:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        # Auto-adjust column widths
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = adjusted_width

    # Quay lại đầu file và trả về dữ liệu Excel dưới dạng response
    output.seek(0)
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=report_top_dishes.xlsx"},
    )

@report.route("/pdf_topdishes")
@role_required(["Quản lý"])
def report_pdf_dishes():
    # Lấy tham số từ URL (GET request)
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    limit_number = request.args.get("limit_number", default=10, type=int)

    # Kiểm tra tham số đầu vào
    if not start_date or not end_date:
        return (
            "Missing 'start_date' or 'end_date' parameter. Use format YYYY-MM-DD.",
            400,
        )

    # Lấy dữ liệu món ăn bán chạy
    dishes = get_dishes_data(start_date, end_date, limit_number)

    if not dishes:
        return "No data found for the given date range.", 404

    # Tạo đối tượng In PDF (canvas) từ ReportLab
    buffer = io.BytesIO()  # Để ghi file PDF vào bộ nhớ
    c = canvas.Canvas(buffer, pagesize=letter)
    font_path = os.path.join(
        os.getcwd(), "website", "static", "fonts", "arial-unicode-ms-regular.ttf"
    )  # Đường dẫn đến font trong thư mục 'static/fonts'
    pdfmetrics.registerFont(TTFont("VietFont", font_path))
    c.setFont("VietFont", 12)

    # Thông tin tiêu đề và ngày in
    current_date = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    user = NhanVien.query.filter_by(idNguoiDung=current_user.MaND).first()
    printed_by = user.getHoTenNV() if user else "Unknown User"

    # In tiêu đề và thông tin bổ sung
    c.setFont("VietFont", 18)
    c.drawString(200, 770, "Báo Cáo Món Ăn Bán Chạy")
    c.setFont("VietFont", 12)
    c.drawString(200, 740, f"Từ: {start_date} Đến: {end_date}")
    c.drawString(200, 725, f"Ngày in: {current_date}")
    c.drawString(200, 710, f"Người in: {printed_by}")

    # Tạo bảng dữ liệu
    y_position = 500
    x_position = 100
    table_data = [["Món Ăn", "Số Lượng Bán"]]  # Tiêu đề bảng

    for dish in dishes:
        table_data.append([dish["DishName"], str(dish["TotalQuantity"])])

    # Tạo bảng với SimpleDocTemplate và Table
    table_style = TableStyle(
        [
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.toColor("#FFD700"),
            ),  # Màu vàng cho tiêu đề
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),  # Màu chữ trắng cho tiêu đề
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),  # Căn giữa tất cả dữ liệu
            ("FONTNAME", (0, 0), (-1, -1), "VietFont"),  # Sử dụng phông chữ VietFont
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("TOPPADDING", (0, 0), (-1, 0), 12),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),  # Lưới của bảng
            ("FONTNAME", (0, 1), (-1, -1), "VietFont"),
        ]
    )

    # Tạo bảng
    table = Table(table_data, colWidths=[200, 200, 200], rowHeights=25)
    table.setStyle(table_style)

    # Đưa bảng vào vị trí y_position
    table.wrapOn(c, x_position, 600)
    table.drawOn(c, x_position, y_position)

    # Kết thúc việc tạo PDF
    c.showPage()
    c.save()

    # Lấy nội dung PDF từ bộ nhớ
    pdf = buffer.getvalue()
    buffer.close()

    # Trả về file PDF dưới dạng response
    return Response(
        pdf,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=report_top_dishes.pdf"
        },
    )

###report doanh thu


def get_revenue_data(start_date, end_date):

    # Chuyển đổi ngày từ định dạng dd-mm-yyyy sang yyyy-mm-dd
    start_date = convert_date_format(start_date)
    end_date = convert_date_format(end_date)

    # Thực thi stored procedure với các tham số đã chuyển đổi
    result = db.session.execute(
        text(
            """
            CALL TinhDoanhThuTheoThang(:start_date, :end_date)
            """
        ),
        {
            "start_date": start_date,
            "end_date": end_date,
        },
    )

    rows = result.fetchall()  # Đọc hết dữ liệu từ câu lệnh

    # Kiểm tra dữ liệu
    print(rows)  # In ra kết quả để kiểm tra cấu trúc dữ liệu (rows)

    # Chuyển đổi kết quả thành danh sách dict
    if rows:
        columns = result.keys()  # Lấy tên các cột từ kết quả SQL
        return [
            dict(zip(columns, row)) for row in rows
        ]  # Kết hợp cột và dữ liệu thành dict

    return []  # Trả về danh sách rỗng nếu không có dữ liệu


@report.route("/revenue", methods=["GET", "POST"])
@role_required(["Quản lý"])
def report_page_revenue():
    if request.method == "POST":
        # Lấy thông tin từ form

        start_date = request.form.get("start_date")
        end_date = request.form.get("end_date")
        if not start_date or not end_date:
            flash("Vui lòng nhập đầy đủ thông tin", "danger")
            return render_template("/admin/report/form_monthrevenue.html")
        try:
            # Chuyển đổi start_date và end_date sang định dạng datetime
            start_date_obj = datetime.strptime(start_date, "%d-%m-%Y")
            end_date_obj = datetime.strptime(end_date, "%d-%m-%Y")
        except ValueError:
            flash("Ngày tháng không hợp lệ. Vui lòng nhập theo định dạng dd-mm-yyyy.", "danger")
            return render_template("/admin/report/form_monthrevenue.html")
        if start_date_obj > end_date_obj:
            flash("Ngày bắt đầu không thể lớn hơn ngày kết thúc", "danger")
            return render_template("/admin/report/form_monthrevenue.html")
        # Gọi hàm lấy dữ liệu
        revenue = get_revenue_data(start_date, end_date)
        print(revenue)
        return render_template(
            "/admin/report/report_monthrevenue.html",
            revenue=revenue,
            start_date=start_date,
            end_date=end_date,
        )

    return render_template("/admin/report/form_monthrevenue.html")


@report.route("/pdf_monthlyrevenue")
@role_required(["Quản lý"])
def report_pdf_revenue():
    # Lấy tham số từ URL (GET request)
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    # Kiểm tra tham số đầu vào
    if not start_date or not end_date:
        return (
            "Missing 'start_date' or 'end_date' parameter. Use format YYYY-MM-DD.",
            400,
        )

    # Lấy dữ liệu doanh thu
    revenue = get_revenue_data(start_date, end_date)

    if not revenue:
        return "No data found for the given date range.", 404

    # Tạo đối tượng In PDF (canvas) từ ReportLab
    buffer = io.BytesIO()  # Để ghi file PDF vào bộ nhớ
    c = canvas.Canvas(buffer, pagesize=letter)
    font_path = os.path.join(
        os.getcwd(), "website", "static", "fonts", "arial-unicode-ms-regular.ttf"
    )  # Đường dẫn đến font trong thư mục 'static/fonts'
    pdfmetrics.registerFont(TTFont("VietFont", font_path))
    c.setFont("VietFont", 12)

    # Thông tin tiêu đề và ngày in
    current_date = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    user = NhanVien.query.filter_by(idNguoiDung=current_user.MaND).first()
    printed_by = user.getHoTenNV() if user else "Unknown User"

    # In tiêu đề và thông tin bổ sung
    c.setFont("VietFont", 18)
    c.drawString(200, 770, "Báo Cáo Doanh Thu Tháng")
    c.setFont("VietFont", 12)
    c.drawString(200, 740, f"Từ: {start_date} Đến: {end_date}")
    c.drawString(200, 725, f"Ngày in: {current_date}")
    c.drawString(200, 710, f"Người in: {printed_by}")

    # Tạo bảng dữ liệu
    y_position = 320
    x_position = 100
    table_data = [["Tháng/Năm", "Doanh Thu Tháng"]]  # Tiêu đề bảng

    for dish in revenue:
        table_data.append([dish["YearMonth"], str(dish["MonthlyRevenue"])])

    # Tạo bảng với SimpleDocTemplate và Table
    table_style = TableStyle(
        [
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.toColor("#FFD700"),
            ),  # Màu vàng cho tiêu đề
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),  # Màu chữ trắng cho tiêu đề
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),  # Căn giữa tất cả dữ liệu
            ("FONTNAME", (0, 0), (-1, -1), "VietFont"),  # Sử dụng phông chữ VietFont
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("TOPPADDING", (0, 0), (-1, 0), 12),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),  # Lưới của bảng
            ("FONTNAME", (0, 1), (-1, -1), "VietFont"),
        ]
    )

    # Tạo bảng
    table = Table(table_data, colWidths=[200, 200], rowHeights=25)
    table.setStyle(table_style)

    # Đưa bảng vào vị trí y_position
    table.wrapOn(c, x_position, 600)
    table.drawOn(c, x_position, y_position)

    # Kết thúc việc tạo PDF
    c.showPage()
    c.save()

    # Lấy nội dung PDF từ bộ nhớ
    pdf = buffer.getvalue()
    buffer.close()

    # Trả về file PDF dưới dạng response
    return Response(
        pdf,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=report_monthly_revenue.pdf"
        },
    )


@report.route("/excel_monthlyrevenue")
@role_required(["Quản lý"])
def report_excel_revenue():
    # Lấy tham số từ URL (GET request)
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    # Kiểm tra tham số đầu vào
    if not start_date or not end_date:
        return (
            "Missing 'start_date' or 'end_date' parameter. Use format YYYY-MM-DD.",
            400,
        )
    # Kiểm tra logic thời gian (nếu cần)
    if start_date > end_date:
        return "'start_date' cannot be later than 'end_date'.", 400

    # Lấy dữ liệu doanh thu
    revenue = get_revenue_data(start_date, end_date)
    if not revenue:
        return "No data found for the given date range.", 404

    # Tạo DataFrame từ kết quả
    df = pd.DataFrame(revenue)

    # Tạo file Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Ghi dữ liệu vào sheet
        df.to_excel(writer, index=False, sheet_name="Report", startrow=5)

        # Lấy workbook và sheet
        workbook = writer.book
        worksheet = writer.sheets["Report"]

        # Lấy ngày in và thông tin người in
        current_date = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        print("MaND:", current_user.get_id())
        user = NhanVien.query.filter_by(idNguoiDung=current_user.MaND).first()
        print("user", user)
        printed_by = user.getHoTenNV() if user else "Unknown User"

        # Ghi tiêu đề và thông tin bổ sung
        worksheet.cell(row=1, column=1, value="Báo cáo doanh thu tháng").font = Font(
            size=16, bold=True, color="0000FF"
        )
        worksheet.cell(row=2, column=1, value=f"Ngày in: {current_date}").font = Font(
            italic=True
        )
        worksheet.cell(row=3, column=1, value=f"Người in: {printed_by}").font = Font(
            italic=True
        )

        # Trang trí tiêu đề cột
        header_fill = PatternFill(
            start_color="FFD700", end_color="FFD700", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Áp dụng định dạng tiêu đề cột
        for col_num, value in enumerate(df.columns, 1):
            cell = worksheet.cell(row=5, column=col_num, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Định dạng toàn bộ dữ liệu
        for row in worksheet.iter_rows(
            min_row=6, max_row=6 + len(df), min_col=1, max_col=len(df.columns)
        ):
            for cell in row:
                cell.border = thin_border

        # Điều chỉnh độ rộng cột
        for col_num in range(1, len(df.columns) + 1):
            col_letter = worksheet.cell(row=5, column=col_num).column_letter
            worksheet.column_dimensions[col_letter].width = 20

        # Định dạng dữ liệu (căn giữa)
        for row_num in range(6, 6 + len(df)):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Quay lại đầu file và trả về response
    output.seek(0)
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=report_monthly_revenue.xlsx"
        },
    )
